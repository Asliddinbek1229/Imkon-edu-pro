import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


async def export_to_excel(data, headings, filepath, summary_rows=None):
    """
    data         — list of tuples (satırlar)
    headings     — list of strings (ustun sarlavhalari)
    summary_rows — list of tuples; har biri JAMI qatori (None = yo'q)
    """
    wb = openpyxl.Workbook()
    sheet = wb.active

    HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    SUMMARY_FILL = PatternFill("solid", fgColor="D6E4F0")
    SUMMARY_FONT = Font(bold=True, size=11)
    ALT_FILL = PatternFill("solid", fgColor="F2F2F2")
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)

    # ── Sarlavha qatori ──────────────────────────────────────────────────────
    for colno, heading in enumerate(headings, start=1):
        cell = sheet.cell(row=1, column=colno, value=heading)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER

    # ── Ma'lumot qatorlari ───────────────────────────────────────────────────
    for rowno, row in enumerate(data, start=2):
        fill = ALT_FILL if rowno % 2 == 0 else None
        for colno, cell_value in enumerate(row, start=1):
            cell = sheet.cell(row=rowno, column=colno, value=cell_value)
            if fill:
                cell.fill = fill
            cell.alignment = Alignment(vertical="center", wrap_text=False)

    # ── JAMI qatorlari ───────────────────────────────────────────────────────
    if summary_rows:
        # Bo'sh ajratuvchi qator
        empty_row = len(data) + 2
        for colno in range(1, len(headings) + 1):
            sheet.cell(row=empty_row, column=colno, value="")

        for offset, summary in enumerate(summary_rows):
            srow = empty_row + 1 + offset
            for colno, cell_value in enumerate(summary, start=1):
                cell = sheet.cell(row=srow, column=colno, value=cell_value)
                cell.font = SUMMARY_FONT
                cell.fill = SUMMARY_FILL
                cell.alignment = Alignment(vertical="center")

    # ── Ustun kengliklarini avtomatik sozlash ────────────────────────────────
    for col_idx in range(1, len(headings) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(headings[col_idx - 1]))
        for rowno in range(2, sheet.max_row + 1):
            val = sheet.cell(row=rowno, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        sheet.column_dimensions[col_letter].width = min(max_len + 4, 40)

    # Sarlavha qatorini muzlatish
    sheet.freeze_panes = "A2"

    wb.save(filepath)
